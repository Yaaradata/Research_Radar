#!/usr/bin/env python3
"""Export bake-off labelling workbook (.xlsx) for human reviewers.

Sheet 1: disagreements (no model outputs shown).
Sheet 2: agreement control (30 random unanimous papers).
Sheet 3: reference enums.
Hidden-style _model_outputs sheet for post-hoc inspection only.
"""

from __future__ import annotations

import argparse
import json
import random
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation

from research_radar.bakeoff import (  # noqa: E402
    APPLICATION_DOMAIN_INSTRUCTION,
    GENERAL_METHOD_VALUE,
    fallback_vocabulary,
    is_general_method,
    load_bakeoff_config,
    vocabulary_from_conn,
)
from research_radar.classification_vocab import AUDIENCE_RELEVANCE
from research_radar.pipeline import connect

LABELLERS = ("subhashini", "urmila", "ranjith")
LABEL_COLS = ("domain", "subdomains", "application_domains", "is_general_method", "reasoning")


def _parse_apps(raw) -> list[str]:
    if raw is None:
        return []
    if isinstance(raw, str):
        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            return [raw]
    return list(raw)


def _candidate_rows(conn, run_id: str) -> dict[int, dict[str, dict]]:
    rows = conn.execute(
        """
        SELECT candidate_id, content_id, domain, subdomains, application_domains,
               primary_audience, ai_relevance
        FROM research_radar.bakeoff_results
        WHERE run_id = %s AND pass_index = 1 AND batch_arrangement = 'A'
        """,
        (run_id,),
    ).fetchall()
    by_cid: dict[int, dict[str, dict]] = {}
    for r in rows:
        cid = int(r["content_id"])
        apps = _parse_apps(r["application_domains"])
        by_cid.setdefault(cid, {})[r["candidate_id"]] = {
            "domain": r.get("domain"),
            "subdomains": r.get("subdomains"),
            "application_domains": apps,
            "is_general_method": is_general_method(apps),
            "primary_audience": r.get("primary_audience"),
            "ai_relevance": r.get("ai_relevance"),
        }
    return by_cid


def _paper_meta(conn, content_ids: list[int]) -> dict[int, dict]:
    if not content_ids:
        return {}
    rows = conn.execute(
        """
        SELECT ci.id AS content_id, ci.title,
               COALESCE(pm.abstract, ci.summary, '') AS abstract
        FROM research_radar.content_items ci
        LEFT JOIN research_radar.paper_metadata pm ON pm.content_id = ci.id
        WHERE ci.id = ANY(%s)
        """,
        (content_ids,),
    ).fetchall()
    return {int(r["content_id"]): dict(r) for r in rows}


def _disagrees(candidate_map: dict[str, dict]) -> bool:
    if len(candidate_map) < 2:
        return False
    domains = {v.get("domain") for v in candidate_map.values()}
    gm_flags = {v.get("is_general_method") for v in candidate_map.values()}
    return len(domains) > 1 or len(gm_flags) > 1


def _add_validation(ws, col_letter: str, options: list[str], start_row: int, end_row: int):
    if not options:
        return
    formula = '"' + ",".join(options) + '"'
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Export bake-off labelling workbook")
    parser.add_argument("--run-id", required=True)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--control-size", type=int, default=30)
    parser.add_argument("--seed", type=int, default=42)
    args = parser.parse_args()

    with connect() as conn:
        vocab = vocabulary_from_conn(conn)
        by_cid = _candidate_rows(conn, args.run_id)
        if not by_cid:
            raise SystemExit(f"No pass-1 results for run_id={args.run_id}")

        disagreements = [cid for cid, cmap in by_cid.items() if _disagrees(cmap)]
        unanimous = [cid for cid, cmap in by_cid.items() if not _disagrees(cmap)]
        rng = random.Random(args.seed)
        control = rng.sample(unanimous, min(args.control_size, len(unanimous)))

        all_ids = disagreements + control
        meta = _paper_meta(conn, all_ids)

    domains = list(vocab.get("domains_list") or fallback_vocabulary()["domains_list"]) + ["Other"]
    applications = list(vocab.get("applications_list") or fallback_vocabulary()["applications_list"])
    if GENERAL_METHOD_VALUE not in applications:
        applications.insert(0, GENERAL_METHOD_VALUE)
    audiences = list(AUDIENCE_RELEVANCE)
    bool_opts = ["TRUE", "FALSE"]

    wb = Workbook()
    ws_ref = wb.active
    ws_ref.title = "Reference"
    ws_ref.append(["field", "allowed_values"])
    ws_ref.append(["domain", ", ".join(domains)])
    ws_ref.append(["application_domains", ", ".join(applications)])
    ws_ref.append(["primary_audience", ", ".join(audiences)])
    ws_ref.append(["is_general_method", "TRUE / FALSE"])
    ws_ref.append([])
    ws_ref.append(["general-method instruction"])
    for line in APPLICATION_DOMAIN_INSTRUCTION.strip().splitlines():
        ws_ref.append([line])

    def _header():
        cols = ["content_id", "title", "abstract"]
        for lab in LABELLERS:
            for c in LABEL_COLS:
                cols.append(f"{lab}_{c}")
        return cols

    def _write_sheet(name: str, cids: list[int]):
        ws = wb.create_sheet(name)
        ws.append(_header())
        for cid in cids:
            m = meta.get(cid, {})
            row = [cid, m.get("title", ""), m.get("abstract", "")]
            for _lab in LABELLERS:
                row.extend(["", "", "", "", ""])
            ws.append(row)
        end = max(2, len(cids) + 1)
        # domain col D for first labeller = col 4; approximate validation on labeller domain cols
        col_idx = 4
        for _lab in LABELLERS:
            col_letter = chr(ord("A") + col_idx - 1) if col_idx <= 26 else "D"
            _add_validation(ws, col_letter, domains, 2, end)
            col_idx += 5
        return ws

    _write_sheet("Disagreements", disagreements)
    _write_sheet("Agreement control", control)

    ws_out = wb.create_sheet("_model_outputs")
    ws_out.append(
        ["content_id", "candidate_id", "domain", "application_domains", "primary_audience", "ai_relevance"]
    )
    for cid, cmap in sorted(by_cid.items()):
        if cid not in all_ids:
            continue
        for cand_id, vals in sorted(cmap.items()):
            ws_out.append(
                [
                    cid,
                    cand_id,
                    vals.get("domain"),
                    json.dumps(vals.get("application_domains") or []),
                    vals.get("primary_audience"),
                    vals.get("ai_relevance"),
                ]
            )

    out = args.output or ROOT / "reports" / f"bakeoff-labels-{args.run_id}.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote {out}")
    print(f"Disagreements: {len(disagreements)}, Agreement control: {len(control)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
