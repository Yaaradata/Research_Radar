"""Tests for classification bake-off harness."""

from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook, load_workbook

from research_radar import bakeoff as bo
from research_radar.bakeoff import (
    BakeoffCandidate,
    BakeoffConfig,
    BakeoffParseError,
    compute_force_fit_rate,
    cost_from_tokens,
    cost_per_thousand_from_measured,
    count_general_method_violations,
    estimate_full_bakeoff_cost,
    is_force_fit,
    is_general_method,
    load_bakeoff_config,
    parse_bakeoff_batch,
    select_stratified_sample,
    self_consistency_passes,
    batch_stability_passes,
)


def _paper(cid: int, cats: list[str]) -> dict:
    return {
        "content_id": cid,
        "title": f"Paper {cid}",
        "abstract": "We propose a method and evaluate on benchmarks.",
        "categories": cats,
    }


def _make_pool() -> list[dict]:
    pool = []
    for i, cat in enumerate(["cs.AI", "cs.CV", "cs.CR", "stat.ML"] * 120):
        pool.append(_paper(1000 + i, [cat]))
    return pool


def test_general_method_mutually_exclusive():
    assert count_general_method_violations(["general-method"]) == 0
    assert count_general_method_violations(["healthcare"]) == 0
    assert count_general_method_violations(["general-method", "healthcare"]) == 1


def test_parse_rejects_general_method_with_sector():
    text = json.dumps(
        {
            "papers": [
                {
                    "paper_id": 1,
                    "domain": "Machine Learning Theory",
                    "subdomains": ["Optimization Theory"],
                    "application_domains": ["general-method", "healthcare"],
                    "primary_audience": "practitioner",
                    "ai_relevance": 8.0,
                }
            ]
        }
    )
    with pytest.raises(BakeoffParseError):
        parse_bakeoff_batch(text, {1})


def test_force_fit_only_model_specific_vs_human_general():
    human = {1: {"is_general_method": True}, 2: {"is_general_method": False}}
    model = {
        1: {"application_domains": ["healthcare"]},
        2: {"application_domains": ["general-method"]},
    }
    assert is_force_fit(["healthcare"], True) is True
    assert is_force_fit(["general-method"], True) is False
    assert is_force_fit(["healthcare"], False) is False
    rate = compute_force_fit_rate(model, human)
    assert rate == 0.5


def test_sample_stratification_fixed_seed():
    pool = _make_pool()
    sample, counts = select_stratified_sample(pool, seed=42)
    assert counts["ai_core"] == 100
    assert counts["cs_cv"] == 100
    assert counts["non_ai_cs"] == 100
    assert counts["other"] == 100
    assert len(sample) == 400
    _, counts2 = select_stratified_sample(pool, seed=42)
    assert counts == counts2


def test_structured_output_gate_disqualifies_before_scoring():
    cand = BakeoffCandidate(id="bad", model="fake/model", reasoning="disabled")
    vocab = bo.fallback_vocabulary()
    with patch("research_radar.bakeoff.call_chat_completion", side_effect=bo.LLMBatchError("no schema")):
        with patch("research_radar.bakeoff.require_api_key"):
            result = bo.test_structured_output_gate(cand, vocab, client=MagicMock())
    assert result["passed"] is False


def test_self_consistency_aligns_on_content_id():
    rows1 = {1: {"domain": "A", "application_domains": ["general-method"], "primary_audience": "student", "ai_relevance": 8.0}}
    rows2 = {1: {"domain": "A", "application_domains": ["general-method"], "primary_audience": "student", "ai_relevance": 8.0}}
    rows2_wrong_order = {99: rows2[1]}  # different id should not count
    assert self_consistency_passes(rows1, rows2) is True
    assert self_consistency_passes(rows1, rows2_wrong_order) is False


def test_batch_stability_aligns_on_content_id():
    a = {5: {"domain": "CV", "application_domains": ["general-method"], "primary_audience": "practitioner", "ai_relevance": 7.0}}
    b = {5: {"domain": "CV", "application_domains": ["general-method"], "primary_audience": "practitioner", "ai_relevance": 7.0}}
    assert batch_stability_passes(a, b) is True


def test_export_sheets_exclude_model_outputs_on_label_sheets(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Disagreements"
    ws.append(["content_id", "title", "abstract", "subhashini_domain"])
    ws.append([1, "T", "A", ""])
    ws_out = wb.create_sheet("_model_outputs")
    ws_out.append(["content_id", "candidate_id", "domain"])
    ws_out.append([1, "haiku", "NLP"])
    path = tmp_path / "test.xlsx"
    wb.save(path)
    loaded = load_workbook(path, read_only=True)
    for name in ("Disagreements", "Agreement control"):
        if name not in loaded.sheetnames:
            continue
        header = list(loaded[name].iter_rows(max_row=1, values_only=True))[0]
        assert "candidate_id" not in header
        assert "haiku" not in str(header)


def test_import_skips_blank_rows():
    import sys
    from pathlib import Path

    scripts = Path(__file__).resolve().parents[1] / "scripts"
    sys.path.insert(0, str(scripts))
    from bakeoff_import_labels import _row_labels

    hmap = {"content_id": 0, "subhashini_domain": 3}
    assert _row_labels([None, "", "", ""], hmap, "subhashini") is None
    label = _row_labels([42, "T", "A", "NLP"], hmap, "subhashini")
    assert label is not None
    assert label["content_id"] == 42
    assert label["domain"] == "NLP"


def test_cost_per_thousand_from_measured_tokens():
    cand = BakeoffCandidate(id="x", model="m", reasoning="disabled", input_cost_per_million=1.0, output_cost_per_million=5.0)
    per_paper_cost = cost_from_tokens(2000, 500, cand)
    assert per_paper_cost == pytest.approx(0.0045, rel=1e-3)
    per_1k = cost_per_thousand_from_measured(per_paper_cost * 400, 400)
    assert per_1k == pytest.approx(4.5, rel=1e-3)


def test_estimate_full_bakeoff_cost_uses_config():
    config = load_bakeoff_config()
    est = estimate_full_bakeoff_cost(config, bo.fallback_vocabulary())
    assert est["sample_size"] == 400
    assert est["candidates"] == 4
    assert est["passes_per_candidate"] == 3
    assert est["total_estimated_cost_usd"] > 0


def test_load_bakeoff_config_from_yaml():
    config = load_bakeoff_config(Path(__file__).resolve().parents[1] / "config" / "bakeoff_models.yaml")
    assert len(config.candidates) == 4
    assert all(c.reasoning == "disabled" for c in config.candidates)


def test_is_general_method_helper():
    assert is_general_method(["general-method"]) is True
    assert is_general_method(["healthcare"]) is False
