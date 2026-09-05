#!/usr/bin/env python3
"""Import completed bake-off labelling workbook into bakeoff_labels."""

from __future__ import annotations

import argparse
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "src"))

from openpyxl import load_workbook

from research_radar.bakeoff import GENERAL_METHOD_VALUE  # noqa: E402
from research_radar.pipeline import connect

LABELLERS = ("subhashini", "urmila", "ranjith")
LABEL_COLS = ("domain", "subdomains", "application_domains", "is_general_method", "reasoning")
SHEETS = ("Disagreements", "Agreement control")


def _parse_bool(val) -> bool | None:
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ("true", "1", "yes"):
        return True
    if s in ("false", "0", "no"):
        return False
    return None


def _parse_apps(val) -> list[str] | None:
    if val is None or (isinstance(val, str) and not val.strip()):
        return None
    if isinstance(val, str):
        parts = [p.strip() for p in val.replace(";", ",").split(",") if p.strip()]
        return parts or None
    return [str(val)]


def _header_map(header: list) -> dict[str, int]:
    return {str(h): i for i, h in enumerate(header) if h}


def _row_labels(row: list, hmap: dict[str, int], labeller: str) -> dict | None:
    cid_idx = hmap.get("content_id")
    if cid_idx is None or cid_idx >= len(row) or row[cid_idx] in (None, ""):
        return None
    try:
        cid = int(row[cid_idx])
    except (TypeError, ValueError):
        return None

    def cell(name):
        key = f"{labeller}_{name}"
        idx = hmap.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    domain = cell("domain")
    if domain is not None:
        domain = str(domain).strip() or None
    subdomains_raw = cell("subdomains")
    subdomains = None
    if subdomains_raw not in (None, ""):
        subdomains = [s.strip() for s in str(subdomains_raw).replace(";", ",").split(",") if s.strip()]
    apps = _parse_apps(cell("application_domains"))
    is_gm = _parse_bool(cell("is_general_method"))
    if is_gm is None and apps:
        is_gm = GENERAL_METHOD_VALUE in apps
    reasoning = cell("reasoning")
    if reasoning is not None:
        reasoning = str(reasoning).strip() or None

    if not any([domain, subdomains, apps, is_gm is not None, reasoning]):
        return None

    return {
        "content_id": cid,
        "labeller": labeller,
        "domain": domain,
        "subdomains": subdomains,
        "application_domains": apps,
        "is_general_method": is_gm,
        "reasoning": reasoning,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Import bake-off labels from xlsx")
    parser.add_argument("--run-id", required=True)
    parser.add_argument("--workbook", type=Path, required=True)
    args = parser.parse_args()

    wb = load_workbook(args.workbook, read_only=True, data_only=True)
    imported = 0
    skipped = 0

    with connect() as conn:
        for sheet_name in SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            hmap = _header_map(list(rows[0]))
            for row in rows[1:]:
                if not row or row[0] in (None, ""):
                    skipped += 1
                    continue
                for lab in LABELLERS:
                    label = _row_labels(list(row), hmap, lab)
                    if label is None:
                        continue
                    conn.execute(
                        """
                        INSERT INTO research_radar.bakeoff_labels (
                            content_id, run_id, labeller, domain, subdomains,
                            application_domains, is_general_method, reasoning
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        ON CONFLICT (content_id, run_id, labeller) DO UPDATE SET
                            domain = EXCLUDED.domain,
                            subdomains = EXCLUDED.subdomains,
                            application_domains = EXCLUDED.application_domains,
                            is_general_method = EXCLUDED.is_general_method,
                            reasoning = EXCLUDED.reasoning,
                            labelled_at = NOW()
                        """,
                        (
                            label["content_id"],
                            args.run_id,
                            label["labeller"],
                            label["domain"],
                            label["subdomains"],
                            label["application_domains"],
                            label["is_general_method"],
                            label["reasoning"],
                        ),
                    )
                    imported += 1
        conn.commit()

    print(f"Imported {imported} label rows, skipped {skipped} blank rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
